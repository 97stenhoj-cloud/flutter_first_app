#!/usr/bin/env python3
"""
Generate INSERT statements from the translated Excel file.
Creates SQL to insert all questions with complete translations.
"""
import openpyxl
import json
import sys

# Enable line buffering for immediate output
sys.stdout.reconfigure(line_buffering=True)

print("Reading Excel file...", flush=True)
wb = openpyxl.load_workbook('/Users/stenhoej/Downloads/Talk Card (6).xlsx')
ws = wb['Sheet v4.0']

# Language mapping
lang_map = {
    'Question (EN)': 'en',
    'DK': 'da',
    'DE': 'de',
    'ES': 'es',
    'PT': 'pt',
    'RO': 'ro',
    'FR': 'fr',
    'IT': 'it',
    'FI': 'fi',
    'NL': 'nl',
    'NO': 'no',
    'SV': 'sv',
    'PL': 'pl'
}

# Find column indices
header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
col_indices = {}
for idx, header in enumerate(header_row):
    if header in lang_map:
        col_indices[lang_map[header]] = idx
    elif header == 'Card Deck':
        col_indices['category'] = idx

print(f"Found columns: {list(col_indices.keys())}", flush=True)

# Collect all questions
questions = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if not row[col_indices.get('category', 0)]:
        continue

    category_name = row[col_indices['category']]

    # Build translations object
    translations = {}
    for lang_code, col_idx in col_indices.items():
        if lang_code in ['category']:
            continue
        text = row[col_idx]
        if text:
            translations[lang_code] = str(text).strip()

    # Skip if no English translation
    if 'en' not in translations:
        continue

    # Determine flags based on category
    is_couple = category_name in ['Love Talks', 'Deep Talks', 'After Dark Talks']
    is_friends = True  # All questions available for friends
    is_family = category_name not in ['After Dark Talks', 'Love Talks']

    questions.append({
        'category_name': category_name,
        'translations': translations,
        'is_couple': is_couple,
        'is_friends': is_friends,
        'is_family': is_family
    })

print(f"\nProcessed {len(questions)} questions", flush=True)
print(f"Categories: {set(q['category_name'] for q in questions)}", flush=True)

# Generate INSERT statements
print("\nGenerating INSERT statements...", flush=True)

# Create batches of 50 inserts
batch_size = 50
total_batches = (len(questions) + batch_size - 1) // batch_size

for batch_num in range(total_batches):
    start_idx = batch_num * batch_size
    end_idx = min(start_idx + batch_size, len(questions))
    batch = questions[start_idx:end_idx]

    filename = f'insert_batch_{batch_num+1:03d}.sql'

    with open(filename, 'w', encoding='utf-8') as f:
        for q in batch:
            # Escape single quotes in JSON by using double dollar quoting
            translations_json = json.dumps(q['translations'], ensure_ascii=False)
            # Escape single quotes for PostgreSQL
            translations_json_escaped = translations_json.replace("'", "''")

            category_escaped = q['category_name'].replace("'", "''")

            sql = f"""INSERT INTO questions (category_name, translations, is_couple, is_friends, is_family)
VALUES ('{category_escaped}', '{translations_json_escaped}'::jsonb, {str(q['is_couple']).lower()}, {str(q['is_friends']).lower()}, {str(q['is_family']).lower()});
"""
            f.write(sql)

    print(f"Created {filename} with {len(batch)} inserts", flush=True)

# Create combined file
print("\nCreating combined INSERT file...", flush=True)
with open('insert_all_questions.sql', 'w', encoding='utf-8') as outfile:
    for batch_num in range(total_batches):
        filename = f'insert_batch_{batch_num+1:03d}.sql'
        with open(filename, 'r', encoding='utf-8') as infile:
            outfile.write(infile.read())

print(f"\nComplete! Created insert_all_questions.sql with {len(questions)} INSERT statements", flush=True)
print(f"Total batches: {total_batches}", flush=True)
print("\nReady to insert into Supabase!", flush=True)
