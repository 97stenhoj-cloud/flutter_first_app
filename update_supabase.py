#!/usr/bin/env python3
"""
Script to update Supabase questions database from translated Excel file.
This script will:
1. Read all questions from the Excel file (Sheet v4.0)
2. Match them to existing questions in Supabase by English text and category
3. Update the translations JSONB field with all languages
4. Handle questions that might be new or different
"""

import openpyxl
import json
import os
import sys

# Configuration
EXCEL_PATH = '/Users/stenhoej/Downloads/Talk Card (6).xlsx'
SHEET_NAME = 'Sheet v4.0'

# Language mapping
LANG_COLUMNS = {
    'en': 3,
    'da': 4,
    'de': 5,
    'es': 6,
    'pt': 7,
    'ro': 8,
    'fr': 9,
    'it': 10,
    'fi': 11,
    'nl': 12,
    'no': 13,
    'sv': 14,
    'pl': 15
}

def load_excel_data():
    """Load all questions from Excel file"""
    print("Loading Excel file...", flush=True)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    questions_by_category = {}
    total_loaded = 0

    for row in range(2, ws.max_row + 1):
        deck_name = ws.cell(row, 1).value

        # Skip invalid rows
        if not deck_name or (isinstance(deck_name, str) and deck_name.startswith('=')):
            continue

        # Get English text
        english_text = ws.cell(row, LANG_COLUMNS['en']).value
        if not english_text:
            continue

        # Build translations dict
        translations = {}
        for lang_code, col_idx in LANG_COLUMNS.items():
            text = ws.cell(row, col_idx).value
            if text and text != '':
                translations[lang_code] = text

        # Store by category
        if deck_name not in questions_by_category:
            questions_by_category[deck_name] = []

        questions_by_category[deck_name].append({
            'english_text': english_text,
            'translations': translations,
            'row': row
        })
        total_loaded += 1

    print(f"Loaded {total_loaded} questions from {len(questions_by_category)} categories", flush=True)
    for category, questions in sorted(questions_by_category.items()):
        print(f"  {category}: {len(questions)} questions", flush=True)

    return questions_by_category

def generate_update_sql(questions_by_category):
    """Generate SQL statements to update Supabase"""

    print("\nGenerating SQL update statements...", flush=True)

    sql_statements = []
    sql_statements.append("-- Update questions with translated text")
    sql_statements.append("-- Generated from Excel file: Talk Card (6).xlsx\n")

    for category_name, questions in sorted(questions_by_category.items()):
        sql_statements.append(f"\n-- Category: {category_name} ({len(questions)} questions)")

        for q in questions:
            english_text = q['english_text'].replace("'", "''")  # Escape single quotes
            translations_json = json.dumps(q['translations'])
            translations_json_escaped = translations_json.replace("'", "''")

            # Update query - match by category_name and English text
            sql = f"""
UPDATE questions
SET translations = '{translations_json_escaped}'::jsonb,
    updated_at = NOW()
WHERE category_name = '{category_name.replace("'", "''")}'
  AND translations->>'en' = '{english_text}';
"""
            sql_statements.append(sql.strip())

    return "\n".join(sql_statements)

def save_sql_file(sql_content):
    """Save SQL to file"""
    output_file = '/Users/stenhoej/Desktop/flutter_first_app/update_questions.sql'

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(sql_content)

    print(f"\nSQL file saved: {output_file}", flush=True)
    return output_file

def generate_verification_sql():
    """Generate SQL to verify the updates"""
    return """
-- Verification queries

-- 1. Check total questions per category
SELECT category_name, COUNT(*) as question_count
FROM questions
GROUP BY category_name
ORDER BY category_name;

-- 2. Check translation completeness (should all have 13 languages)
SELECT
    category_name,
    COUNT(*) as total_questions,
    SUM(CASE WHEN translations ? 'en' THEN 1 ELSE 0 END) as has_en,
    SUM(CASE WHEN translations ? 'da' THEN 1 ELSE 0 END) as has_da,
    SUM(CASE WHEN translations ? 'de' THEN 1 ELSE 0 END) as has_de,
    SUM(CASE WHEN translations ? 'es' THEN 1 ELSE 0 END) as has_es,
    SUM(CASE WHEN translations ? 'fi' THEN 1 ELSE 0 END) as has_fi,
    SUM(CASE WHEN translations ? 'fr' THEN 1 ELSE 0 END) as has_fr,
    SUM(CASE WHEN translations ? 'it' THEN 1 ELSE 0 END) as has_it,
    SUM(CASE WHEN translations ? 'nl' THEN 1 ELSE 0 END) as has_nl,
    SUM(CASE WHEN translations ? 'no' THEN 1 ELSE 0 END) as has_no,
    SUM(CASE WHEN translations ? 'pl' THEN 1 ELSE 0 END) as has_pl,
    SUM(CASE WHEN translations ? 'pt' THEN 1 ELSE 0 END) as has_pt,
    SUM(CASE WHEN translations ? 'ro' THEN 1 ELSE 0 END) as has_ro,
    SUM(CASE WHEN translations ? 'sv' THEN 1 ELSE 0 END) as has_sv
FROM questions
GROUP BY category_name
ORDER BY category_name;

-- 3. Sample questions to verify translations look correct
SELECT
    category_name,
    translations->>'en' as english,
    translations->>'fi' as finnish,
    translations->>'sv' as swedish
FROM questions
WHERE category_name = 'Love Talks'
LIMIT 5;
"""

def main():
    sys.stdout.reconfigure(line_buffering=True)

    print("="*60, flush=True)
    print("Supabase Question Update Script", flush=True)
    print("="*60, flush=True)

    # Load Excel data
    questions_by_category = load_excel_data()

    # Generate SQL
    sql_content = generate_update_sql(questions_by_category)

    # Add verification queries
    sql_content += "\n\n" + generate_verification_sql()

    # Save to file
    sql_file = save_sql_file(sql_content)

    # Statistics
    total_questions = sum(len(questions) for questions in questions_by_category.values())
    total_categories = len(questions_by_category)

    print("\n" + "="*60, flush=True)
    print("Summary:", flush=True)
    print(f"  Total questions: {total_questions}", flush=True)
    print(f"  Total categories: {total_categories}", flush=True)
    print(f"  SQL file: {sql_file}", flush=True)
    print("\nNext steps:", flush=True)
    print("  1. Review the SQL file to ensure it looks correct", flush=True)
    print("  2. Apply the updates to Supabase using the MCP tools", flush=True)
    print("  3. Run the verification queries to check the results", flush=True)
    print("="*60, flush=True)

if __name__ == "__main__":
    main()
