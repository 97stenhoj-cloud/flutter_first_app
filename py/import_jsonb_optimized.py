"""
Import Talk Card questions to Supabase with JSONB translations
OPTIMAL FORMAT: One row per question with all translations in JSONB
"""

import openpyxl
from supabase import create_client, Client

# Your Supabase credentials
SUPABASE_URL = "https://tpjsebutbieghpmvpktv.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRwanNlYnV0YmllZ2hwbXZwa3R2Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc1OTY0NTgwOSwiZXhwIjoyMDc1MjIxODA5fQ.fUb9e9K7zT-2PlPtSsY5aoGG8xLRVSudLtTBzZa4e-I"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Language mapping: column index -> language code
LANGUAGES = {
    2: 'en',  # Column C - English
    3: 'da',  # Column D - Danish
    4: 'de',  # Column E - German
    5: 'es',  # Column F - Spanish
    6: 'pt',  # Column G - Portuguese
    7: 'ro',  # Column H - Romanian
    8: 'fr',  # Column I - French
}

def map_deck_to_game_mode(deck_name: str) -> dict:
    """Map deck name to boolean flags based on exact categorization"""
    
    deck_normalized = deck_name.strip()
    
    couple_decks = [
        'Love Talks',
        'Deep Talks',
        'Silly Talks',
        'Car Talks',
        'Spicy Talks',
        'Do-you-dare Talks',
        'Love Languages Remix Talks'
    ]
    
    friends_decks = [
        'Cozy Talks',
        'Party Night Talks',
        'Silly Talks',
        'Car Talks',
        'Unpopular Opinions XL',
        'Plot Twists & Dilemmas',
        'After Dark Talks'
    ]
    
    family_decks = [
        'Cozy Talks',
        'Car Talks',
        'History Talks',
        'Silly Talks',
        'Would You Rather Talks',
        'Tiny Talks',
        'The Good Old Days Talks'
    ]
    
    is_couple = deck_normalized in couple_decks
    is_friends = deck_normalized in friends_decks
    is_family = deck_normalized in family_decks
    
    if not (is_couple or is_friends or is_family):
        print(f"  ⚠️  Warning: '{deck_normalized}' not in any category, defaulting to Friends")
        is_friends = True
    
    return {
        'is_couple': is_couple,
        'is_friends': is_friends,
        'is_family': is_family
    }

def import_questions(excel_file: str):
    """Import questions with JSONB translations"""
    
    print("📖 Loading Excel file...")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Sheet v2.0']
    
    print("🌍 Importing questions with JSONB translations...")
    print("📦 Format: One row per question, all languages in JSONB\n")
    
    questions_batch = []
    total_imported = 0
    skipped = 0
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        deck_name = row[0]  # Column A: Card Deck
        
        if not deck_name:
            continue
        
        # Build translations object with all languages
        translations = {}
        has_any_translation = False
        
        for col_idx, lang_code in LANGUAGES.items():
            question_text = row[col_idx]
            
            if question_text and str(question_text).strip():
                translations[lang_code] = str(question_text).strip()
                has_any_translation = True
        
        # Skip if no translations found
        if not has_any_translation:
            skipped += 1
            continue
        
        # Get game mode flags
        game_mode_flags = map_deck_to_game_mode(deck_name)
        
        # Create question object
        questions_batch.append({
            'category_name': deck_name,
            'translations': translations,
            'is_couple': game_mode_flags['is_couple'],
            'is_friends': game_mode_flags['is_friends'],
            'is_family': game_mode_flags['is_family'],
        })
        
        # Insert in batches of 50 (smaller batches for JSONB)
        if len(questions_batch) >= 50:
            try:
                supabase.table('questions').insert(questions_batch).execute()
                total_imported += len(questions_batch)
                print(f"  ✅ Imported {total_imported} questions...")
                questions_batch = []
            except Exception as e:
                print(f"  ❌ Error at row {row_num}: {e}")
                return
    
    # Insert remaining questions
    if questions_batch:
        try:
            supabase.table('questions').insert(questions_batch).execute()
            total_imported += len(questions_batch)
            print(f"  ✅ Imported final {len(questions_batch)} questions")
        except Exception as e:
            print(f"  ❌ Error: {e}")
            return
    
    print(f"\n✨ Done!")
    print(f"📊 Total questions imported: {total_imported}")
    print(f"⚠️  Skipped rows: {skipped}")
    print(f"\n💡 Each question contains translations for: {', '.join(LANGUAGES.values())}")
    print(f"📈 Expected: ~1,200 questions (one per unique question)")

if __name__ == "__main__":
    print("=" * 60)
    print("Talk Card Questions Import - JSONB Optimized")
    print("=" * 60)
    
    excel_file = "/Users/stenhoej/Desktop/flutter_first_app/py/Talk_Card__2_.xlsx"
    import_questions(excel_file)
