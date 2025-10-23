"""
Import Talk Card questions to Supabase (English only)
Matches your actual table structure with is_couple, is_friends, is_family
"""

import openpyxl
from supabase import create_client, Client

# Your Supabase credentials
SUPABASE_URL = "https://tpjsebutbieghpmvpktv.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InRwanNlYnV0YmllZ2hwbXZwa3R2Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc1OTY0NTgwOSwiZXhwIjoyMDc1MjIxODA5fQ.fUb9e9K7zT-2PlPtSsY5aoGG8xLRVSudLtTBzZa4e-I"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

def map_deck_to_game_mode(deck_name: str) -> dict:
    """Map deck name to boolean flags based on exact categorization"""
    
    # Normalize the deck name for comparison
    deck_normalized = deck_name.strip()
    
    # Define exact mappings
    couple_decks = [
        'Love Talks',
        'Deep Talks',
        'Silly Talks',
        'Car Talks',
        'Spicy Talks',
        'Do-you-dare Talks',
        'Love Languages Remix Talks'
    ]
    
    friends_decks = [
        'Cozy Talks',
        'Party Night Talks',
        'Silly Talks',
        'Car Talks',
        'Unpopular Opinions XL',
        'Plot Twists & Dilemmas',
        'After Dark Talks'
    ]
    
    family_decks = [
        'Cozy Talks',
        'Car Talks',
        'History Talks',  # Assuming this is "Family Legends Talks"
        'Silly Talks',
        'Would You Rather Talks',
        'Tiny Talks',
        'The Good Old Days Talks'
    ]
    
    # Check which categories this deck belongs to
    is_couple = deck_normalized in couple_decks
    is_friends = deck_normalized in friends_decks
    is_family = deck_normalized in family_decks
    
    # If not found in any category, default to friends
    if not (is_couple or is_friends or is_family):
        print(f"  ⚠️  Warning: '{deck_normalized}' not in any category, defaulting to Friends")
        is_friends = True
    
    return {
        'is_couple': is_couple,
        'is_friends': is_friends,
        'is_family': is_family
    }

def import_questions(excel_file: str):
    """Import English questions from Excel to Supabase"""
    
    print("📖 Loading Excel file...")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Sheet v2.0']
    
    print("🌍 Importing English questions...")
    
    questions_batch = []
    total_imported = 0
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        deck_name = row[0]  # Column A: Card Deck
        question_text = row[2]  # Column C: Question (EN)
        
        if not deck_name or not question_text:
            continue
        
        game_mode_flags = map_deck_to_game_mode(deck_name)
        
        questions_batch.append({
            'category_name': deck_name,
            'text': question_text,
            'is_couple': game_mode_flags['is_couple'],
            'is_friends': game_mode_flags['is_friends'],
            'is_family': game_mode_flags['is_family'],
        })
        
        # Insert in batches of 100
        if len(questions_batch) >= 100:
            try:
                supabase.table('questions').insert(questions_batch).execute()
                total_imported += len(questions_batch)
                print(f"  ✅ Imported {total_imported} questions...")
                questions_batch = []
            except Exception as e:
                print(f"  ❌ Error: {e}")
                return
    
    # Insert remaining questions
    if questions_batch:
        try:
            supabase.table('questions').insert(questions_batch).execute()
            total_imported += len(questions_batch)
            print(f"  ✅ Imported {len(questions_batch)} more questions")
        except Exception as e:
            print(f"  ❌ Error: {e}")
            return
    
    print(f"\n✨ Done! Total questions imported: {total_imported}")
    print(f"📊 Expected: ~1,200 English questions")

if __name__ == "__main__":
    print("=" * 50)
    print("Talk Card Questions Import (English)")
    print("=" * 50)
    
    excel_file = "/Users/stenhoej/Desktop/flutter_first_app/py/Talk_Card__2_.xlsx"
    import_questions(excel_file)